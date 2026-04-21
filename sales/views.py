"""
Sales Views - Reports and Analytics
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, F
from django.db.models.functions import TruncDate, TruncMonth, ExtractHour
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from .models import Sale
from pos.models import POSTransaction, POSTransactionItem, POSSession, POSPayment
from cashregister.models import CashShift, PaymentMethod
from stocks.models import Product, ProductCategory
from decorators.decorators import group_required


@login_required
@group_required(['Admin'])
def reports_dashboard(request):
    """Reports dashboard with overview."""
    today = timezone.localdate()
    month_start = today.replace(day=1)
    
    # Today stats
    today_transactions = POSTransaction.objects.filter(
        created_at__date=today,
        status='completed'
    )
    today_sales = today_transactions.aggregate(total=Sum('total'))['total'] or Decimal('0')
    today_count = today_transactions.count()
    
    # Month stats
    month_transactions = POSTransaction.objects.filter(
        created_at__date__gte=month_start,
        status='completed'
    )
    month_sales = month_transactions.aggregate(total=Sum('total'))['total'] or Decimal('0')
    month_count = month_transactions.count()
    
    # Top products today - using subtotal instead of total
    top_products = POSTransactionItem.objects.filter(
        transaction__created_at__date=today,
        transaction__status='completed'
    ).values('product__name').annotate(
        total_qty=Sum('quantity'),
        total_amount=Sum('subtotal')
    ).order_by('-total_amount')[:5]
    
    # Recent transactions - using cashier instead of opened_by
    recent_transactions = POSTransaction.objects.filter(
        status='completed'
    ).select_related('session__cash_shift__cash_register', 'session__cash_shift__cashier').order_by('-created_at')[:10]
    
    # Sales by payment method today - AHORA FUNCIONAL
    payment_stats = POSPayment.objects.filter(
        transaction__created_at__date=today,
        transaction__status='completed'
    ).values(
        'payment_method__name',
        'payment_method__icon',
        'payment_method__color'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    context = {
        'today_sales': today_sales,
        'today_count': today_count,
        'month_sales': month_sales,
        'month_count': month_count,
        'top_products': top_products,
        'recent_transactions': recent_transactions,
        'payment_stats': payment_stats,
        'today': today,
    }
    return render(request, 'sales/reports_dashboard.html', context)


@login_required
@group_required(['Admin'])
def balance_consolidado(request):
    """Reporte unificado de Ventas vs Gastos con neto operativo.

    Pensado para que el dueño vea de un vistazo cuanta plata le quedo al
    final del periodo sin tener que cruzar dos reportes. Las categorias
    marcadas como inversion (Compras de mercaderia) se muestran aparte y
    NO se restan del neto — a pedido explicito del cliente porque ese
    dinero se convierte en stock, no es gasto puro.
    """
    from expenses.models import Expense
    from stocks.models import StockMovement
    from django.db.models import ExpressionWrapper, DecimalField, Q, Case, When, Value, CharField

    today = timezone.localdate()
    period = (request.GET.get('period') or 'month').lower()
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')

    if period == 'today':
        date_from, date_to = today, today
    elif period == 'week':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'custom' and date_from_str and date_to_str:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    else:
        period = 'month'
        date_from = today.replace(day=1)
        date_to = today

    # Ventas completadas en el rango
    sales_qs = POSTransaction.objects.filter(
        status='completed',
        completed_at__date__gte=date_from,
        completed_at__date__lte=date_to,
    )
    total_sales = sales_qs.aggregate(total=Sum('total'))['total'] or Decimal('0')
    sales_count = sales_qs.count()

    # Gastos del rango, separados entre operativos (restan) e inversion (no)
    expenses_qs = Expense.objects.filter(
        expense_date__gte=date_from,
        expense_date__lte=date_to,
    ).select_related('category')

    operating_expenses = expenses_qs.filter(category__is_investment=False)
    investment_expenses = expenses_qs.filter(category__is_investment=True)

    total_operating = operating_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_investment = investment_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Breakdown por categoria
    operating_by_cat = (
        operating_expenses
        .values('category__name', 'category__color')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    investment_by_cat = (
        investment_expenses
        .values('category__name', 'category__color')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    # Pérdidas de mercadería (valuadas al costo). Unificamos dos fuentes:
    #   1) Ajustes manuales de stock con motivo de merma (movement_type='adjustment_out').
    #   2) Consumo interno cargado desde el POS (movement_type='sale' con reference que
    #      empieza con 'Consumo interno ', ver pos/services.py:733).
    # Excluimos "Corrección de Error" y "Devolución" porque no son pérdidas económicas.
    # Normalizamos todas las referencias del POS al bucket 'Consumo Interno' para que
    # aparezcan junto con los ajustes manuales del mismo motivo.
    LOSS_REASONS = [
        'Robo / Pérdida',
        'Mercadería Dañada',
        'Mercadería Vencida',
        'Consumo Interno',
        'Conteo Físico / Inventario',
    ]

    losses_qs = StockMovement.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).filter(
        Q(movement_type='adjustment_out', reference__in=LOSS_REASONS)
        | Q(movement_type='sale', reference__startswith='Consumo interno ')
    ).select_related('product').annotate(
        loss_reason=Case(
            When(movement_type='sale', then=Value('Consumo Interno')),
            default=F('reference'),
            output_field=CharField(),
        )
    )

    # Pérdida al costo = |quantity| * unit_cost (quantity es negativa en salidas)
    cost_expr = ExpressionWrapper(
        F('quantity') * F('unit_cost') * Decimal('-1'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )

    total_losses = losses_qs.aggregate(
        total=Sum(cost_expr)
    )['total'] or Decimal('0')

    # Agregado por motivo con detalle de productos embebido.
    # Estructura final: [{reference, total_cost, total_qty, movement_count, products: [...]}, ...]
    reason_aggregates = (
        losses_qs.values('loss_reason')
        .annotate(
            total_cost=Sum(cost_expr),
            total_qty=Sum(F('quantity') * Decimal('-1')),
            movement_count=Count('id'),
        )
        .order_by('-total_cost')
    )
    product_rows = (
        losses_qs.values('loss_reason', 'product__id', 'product__name')
        .annotate(
            total_cost=Sum(cost_expr),
            total_qty=Sum(F('quantity') * Decimal('-1')),
            movement_count=Count('id'),
        )
        .order_by('loss_reason', '-total_cost')
    )
    products_by_reason = {}
    for row in product_rows:
        products_by_reason.setdefault(row['loss_reason'], []).append(row)

    losses_by_reason = []
    for agg in reason_aggregates:
        losses_by_reason.append({
            'reference': agg['loss_reason'],
            'total_cost': agg['total_cost'],
            'total_qty': agg['total_qty'],
            'movement_count': agg['movement_count'],
            'products': products_by_reason.get(agg['loss_reason'], []),
        })

    neto_operativo = total_sales - total_operating - total_losses

    context = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'total_sales': total_sales,
        'sales_count': sales_count,
        'total_operating': total_operating,
        'total_investment': total_investment,
        'total_losses': total_losses,
        'operating_by_cat': operating_by_cat,
        'investment_by_cat': investment_by_cat,
        'losses_by_reason': losses_by_reason,
        'neto_operativo': neto_operativo,
    }
    return render(request, 'sales/balance_consolidado.html', context)


@login_required
@group_required(['Admin'])
def sale_list(request):
    """List sales/transactions."""
    transactions = POSTransaction.objects.filter(
        status='completed'
    ).select_related('session__cash_shift__cash_register', 'session__cash_shift__cashier').order_by('-created_at')
    
    # Filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        transactions = transactions.filter(created_at__date__gte=date_from)
    if date_to:
        transactions = transactions.filter(created_at__date__lte=date_to)
    
    total = transactions.aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    context = {
        'transactions': transactions[:100],
        'total': total,
        'count': transactions.count(),
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'sales/sale_list.html', context)


@login_required
@group_required(['Admin'])
def daily_report(request):
    """Daily sales report."""
    date = request.GET.get('date', '')
    if date:
        report_date = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        report_date = timezone.localdate()
    
    transactions = POSTransaction.objects.filter(
        created_at__date=report_date,
        status='completed'
    ).select_related('session__cash_shift__cash_register', 'session__cash_shift__cashier')
    
    # Stats
    stats = transactions.aggregate(
        total_sales=Sum('total'),
        count=Count('id'),
        avg=Avg('total'),
        discount=Sum('discount_total')
    )
    
    # By cashier (through session)
    by_cashier = transactions.values(
        'session__cash_shift__cashier__username',
        'session__cash_shift__cashier__first_name',
        'session__cash_shift__cashier__last_name'
    ).annotate(
        total=Sum('total'),
        count=Count('id')
    )
    
    # By hour - compatible with both SQLite and PostgreSQL
    by_hour = transactions.annotate(
        hour=ExtractHour('created_at')
    ).values('hour').annotate(
        total=Sum('total'),
        count=Count('id')
    ).order_by('hour')
    
    # By payment method
    by_payment = POSPayment.objects.filter(
        transaction__created_at__date=report_date,
        transaction__status='completed'
    ).values(
        'payment_method__name',
        'payment_method__icon',
        'payment_method__color'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    context = {
        'report_date': report_date,
        'transactions': transactions,
        'stats': stats,
        'by_cashier': by_cashier,
        'by_hour': by_hour,
        'by_payment': by_payment,
    }
    return render(request, 'sales/daily_report.html', context)


@login_required
@group_required(['Admin'])
def period_report(request):
    """Period sales report."""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)
    
    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()
    
    transactions = POSTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='completed'
    )
    
    # Stats
    stats = transactions.aggregate(
        total_sales=Sum('total'),
        count=Count('id'),
        avg=Avg('total'),
        discount=Sum('discount_total')
    )
    
    # Daily breakdown
    daily_sales = transactions.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Sum('total'),
        count=Count('id')
    ).order_by('date')
    
    # By payment method
    by_payment = POSPayment.objects.filter(
        transaction__created_at__date__gte=start_date,
        transaction__created_at__date__lte=end_date,
        transaction__status='completed'
    ).values(
        'payment_method__name',
        'payment_method__icon',
        'payment_method__color'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'stats': stats,
        'daily_sales': list(daily_sales),
        'by_payment': by_payment,
    }
    return render(request, 'sales/period_report.html', context)


@login_required
@group_required(['Admin'])
def products_report(request):
    """Products sales report."""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)
    
    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()
    
    items = POSTransactionItem.objects.filter(
        transaction__created_at__date__gte=start_date,
        transaction__created_at__date__lte=end_date,
        transaction__status='completed'
    ).values(
        'product__id',
        'product__name',
        'product__category__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_amount=Sum('subtotal'),
        avg_price=Avg('unit_price')
    ).order_by('-total_amount')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'items': items[:50],
    }
    return render(request, 'sales/products_report.html', context)


@login_required
@group_required(['Admin'])
def categories_report(request):
    """Categories sales report."""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)
    
    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()
    
    items = POSTransactionItem.objects.filter(
        transaction__created_at__date__gte=start_date,
        transaction__created_at__date__lte=end_date,
        transaction__status='completed'
    ).values(
        'product__category__id',
        'product__category__name',
        'product__category__color'
    ).annotate(
        total_qty=Sum('quantity'),
        total_amount=Sum('subtotal'),
        product_count=Count('product', distinct=True)
    ).order_by('-total_amount')
    
    total = sum(item['total_amount'] or 0 for item in items)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'items': items,
        'total': total,
    }
    return render(request, 'sales/categories_report.html', context)


@login_required
@group_required(['Admin'])
def cashiers_report(request):
    """Cashiers performance report."""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)
    
    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()
    
    transactions = POSTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='completed'
    ).values(
        'session__cash_shift__cashier__id',
        'session__cash_shift__cashier__username',
        'session__cash_shift__cashier__first_name',
        'session__cash_shift__cashier__last_name'
    ).annotate(
        total=Sum('total'),
        count=Count('id')
    ).order_by('-total')
    
    # Calculate average per cashier
    for t in transactions:
        if t['count'] > 0:
            t['avg'] = t['total'] / t['count']
        else:
            t['avg'] = Decimal('0')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'cashiers': transactions,
    }
    return render(request, 'sales/cashiers_report.html', context)


@login_required
@group_required(['Admin'])
def export_excel(request):
    """Export sales report to Excel - multi-sheet with full detail."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)

    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()

    # --- fetch data ---
    transactions = POSTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='completed',
    ).select_related(
        'session__cash_shift__cashier',
        'session__cash_shift__cash_register',
    ).prefetch_related('payments__payment_method').order_by('-created_at')

    stats = transactions.aggregate(
        total_sales=Sum('total'),
        count=Count('id'),
        avg=Avg('total'),
        discount=Sum('discount_total'),
    )

    daily_list = list(
        transactions.annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(total=Sum('total'), count=Count('id'))
        .order_by('date')
    )

    by_payment = list(
        POSPayment.objects.filter(
            transaction__created_at__date__gte=start_date,
            transaction__created_at__date__lte=end_date,
            transaction__status='completed',
        )
        .values('payment_method__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    top_products = list(
        POSTransactionItem.objects.filter(
            transaction__created_at__date__gte=start_date,
            transaction__created_at__date__lte=end_date,
            transaction__status='completed',
        )
        .values('product__name', 'product__category__name')
        .annotate(
            total_qty=Sum('quantity'),
            total_amount=Sum('subtotal'),
            avg_price=Avg('unit_price'),
        )
        .order_by('-total_amount')
    )

    by_cashier = list(
        transactions.values(
            'session__cash_shift__cashier__first_name',
            'session__cash_shift__cashier__last_name',
            'session__cash_shift__cashier__username',
        )
        .annotate(total=Sum('total'), count=Count('id'))
        .order_by('-total')
    )

    # --- style helpers ---
    C_PURPLE = '2D1E5F'
    C_PINK   = 'E91E8C'
    C_YELLOW = 'F5D000'
    C_LGRAY  = 'F2F2F2'
    C_WHITE  = 'FFFFFF'

    def fill(c):
        return PatternFill(start_color=c, end_color=c, fill_type='solid')

    def border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def header_row(ws, row_num, headers, color=C_PURPLE):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = Font(bold=True, color=C_WHITE, size=9)
            cell.fill = fill(color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border()
        ws.row_dimensions[row_num].height = 22

    def auto_width(ws, minimum=12):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, minimum)

    def fmt(v):
        return float(v) if v is not None else 0.0

    DAYS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

    wb = openpyxl.Workbook()

    # ============================================================
    # HOJA 1 — RESUMEN DEL PERÍODO
    # ============================================================
    ws1 = wb.active
    ws1.title = 'Resumen'
    from openpyxl.utils import get_column_letter

    ws1.merge_cells('A1:C1')
    c = ws1['A1']
    c.value = 'REPORTE DE VENTAS — CHE GOLOSO'
    c.font = Font(bold=True, size=15, color=C_WHITE)
    c.fill = fill(C_PURPLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:C2')
    c = ws1['A2']
    c.value = (
        f'Período: {start_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}'
        f'  |  Generado: {timezone.localdate().strftime("%d/%m/%Y")}'
    )
    c.font = Font(size=9, color='555555')
    c.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 16

    header_row(ws1, 3, ['Indicador', 'Valor', 'Detalle'], C_PINK)

    resumen_rows = [
        ('Total Vendido',            f'${fmt(stats["total_sales"]):,.2f}',  ''),
        ('Cantidad de Transacciones', stats['count'] or 0,                  ''),
        ('Ticket Promedio',           f'${fmt(stats["avg"]):,.2f}',         ''),
        ('Total Descuentos',          f'${fmt(stats["discount"]):,.2f}',    ''),
        ('Días del Período',          (end_date - start_date).days + 1,     ''),
        ('Promedio Diario',           f'${fmt(stats["total_sales"]) / max((end_date - start_date).days + 1, 1):,.2f}', ''),
    ]

    for i, (label, val, detail) in enumerate(resumen_rows, 4):
        bg = C_LGRAY if i % 2 == 0 else C_WHITE
        for col, v in enumerate([label, val, detail], 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(bold=(col == 1), size=10)
            cell.alignment = Alignment(vertical='center')

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 20

    # ============================================================
    # HOJA 2 — TRANSACCIONES DETALLADAS
    # ============================================================
    ws2 = wb.create_sheet('Transacciones')

    header_row(ws2, 1, [
        'N° Ticket', 'Fecha', 'Hora', 'Caja',
        'Cajero', 'Ítems', 'Subtotal', 'Descuento', 'Total', 'Métodos de Pago',
    ])

    for ri, t in enumerate(transactions, 2):
        cashier  = t.session.cash_shift.cashier       if t.session and t.session.cash_shift else None
        register = t.session.cash_shift.cash_register if t.session and t.session.cash_shift else None
        pmethods = ', '.join(
            p.payment_method.name for p in t.payments.all() if p.payment_method
        ) or '-'

        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        row_vals = [
            t.ticket_number or '',
            t.created_at.strftime('%d/%m/%Y'),
            t.created_at.strftime('%H:%M'),
            register.name if register else '-',
            cashier.get_full_name() if cashier else '-',
            t.items_count,
            fmt(t.subtotal),
            fmt(t.discount_total),
            fmt(t.total),
            pmethods,
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='center')
            if col in (7, 8, 9):
                cell.number_format = '#,##0.00'

    auto_width(ws2)

    # ============================================================
    # HOJA 3 — VENTAS POR DÍA
    # ============================================================
    ws3 = wb.create_sheet('Ventas por Día')
    header_row(ws3, 1, ['Fecha', 'Día de la Semana', 'Transacciones', 'Total del Día', 'Ticket Promedio'])

    for ri, day in enumerate(daily_list, 2):
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        avg_day = fmt(day['total']) / day['count'] if day['count'] else 0.0
        for col, val in enumerate([
            day['date'].strftime('%d/%m/%Y'),
            DAYS_ES[day['date'].weekday()],
            day['count'],
            fmt(day['total']),
            avg_day,
        ], 1):
            cell = ws3.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(size=9)
            if col in (4, 5):
                cell.number_format = '#,##0.00'

    # Total row
    tr = len(daily_list) + 2
    for col, val in enumerate(['TOTAL', '', sum(d['count'] for d in daily_list), fmt(stats['total_sales']), ''], 1):
        cell = ws3.cell(row=tr, column=col, value=val)
        cell.font = Font(bold=True, color=C_WHITE)
        cell.fill = fill(C_YELLOW)
        cell.border = border()
        if col == 4:
            cell.number_format = '#,##0.00'
        # Override text color for yellow fill
        cell.font = Font(bold=True, color='1A1A2E')
    auto_width(ws3)

    # ============================================================
    # HOJA 4 — MÉTODOS DE PAGO
    # ============================================================
    ws4 = wb.create_sheet('Métodos de Pago')
    header_row(ws4, 1, ['Método de Pago', 'Transacciones', 'Total Recaudado', '% del Total'])

    total_v = fmt(stats['total_sales'])
    for ri, pm in enumerate(by_payment, 2):
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        total_pm = fmt(pm['total'])
        pct = (total_pm / total_v * 100) if total_v else 0.0
        for col, val in enumerate([pm['payment_method__name'], pm['count'], total_pm, round(pct, 2)], 1):
            cell = ws4.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(size=9)
            if col == 3:
                cell.number_format = '#,##0.00'
            if col == 4:
                cell.number_format = '0.00'

    auto_width(ws4)

    # ============================================================
    # HOJA 5 — PRODUCTOS VENDIDOS
    # ============================================================
    ws5 = wb.create_sheet('Productos Vendidos')
    header_row(ws5, 1, ['Producto', 'Categoría', 'Cantidad Vendida', 'Ingresos Totales', 'Precio Prom. Venta'])

    for ri, item in enumerate(top_products, 2):
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        for col, val in enumerate([
            item['product__name'],
            item['product__category__name'] or 'Sin categoría',
            float(item['total_qty'] or 0),
            float(item['total_amount'] or 0),
            float(item['avg_price'] or 0),
        ], 1):
            cell = ws5.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(size=9)
            if col in (4, 5):
                cell.number_format = '#,##0.00'
            if col == 3:
                cell.number_format = '0.000'

    auto_width(ws5)

    # ============================================================
    # HOJA 6 — POR CAJERO
    # ============================================================
    ws6 = wb.create_sheet('Por Cajero')
    header_row(ws6, 1, ['Cajero', 'Transacciones', 'Total Vendido', 'Ticket Promedio'])

    for ri, c in enumerate(by_cashier, 2):
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        fn = c['session__cash_shift__cashier__first_name'] or ''
        ln = c['session__cash_shift__cashier__last_name'] or ''
        name = f'{fn} {ln}'.strip() or c['session__cash_shift__cashier__username'] or '-'
        avg_cashier = fmt(c['total']) / c['count'] if c['count'] else 0.0
        for col, val in enumerate([name, c['count'], fmt(c['total']), avg_cashier], 1):
            cell = ws6.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.font = Font(size=9)
            if col in (3, 4):
                cell.number_format = '#,##0.00'

    auto_width(ws6)

    # --- output ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_ventas_{start_date}_{end_date}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@group_required(['Admin'])
def export_pdf(request):
    """Export sales report to PDF using ReportLab."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        start_date = timezone.localdate() - timedelta(days=30)

    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.localdate()

    # --- fetch data ---
    transactions = POSTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='completed',
    )

    stats = transactions.aggregate(
        total_sales=Sum('total'),
        count=Count('id'),
        avg=Avg('total'),
        discount=Sum('discount_total'),
    )

    daily_list = list(
        transactions.annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(total=Sum('total'), count=Count('id'))
        .order_by('date')
    )

    by_payment = list(
        POSPayment.objects.filter(
            transaction__created_at__date__gte=start_date,
            transaction__created_at__date__lte=end_date,
            transaction__status='completed',
        )
        .values('payment_method__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )

    top_products = list(
        POSTransactionItem.objects.filter(
            transaction__created_at__date__gte=start_date,
            transaction__created_at__date__lte=end_date,
            transaction__status='completed',
        )
        .values('product__name', 'product__category__name')
        .annotate(total_qty=Sum('quantity'), total_amount=Sum('subtotal'))
        .order_by('-total_amount')[:25]
    )

    by_cashier = list(
        transactions.values(
            'session__cash_shift__cashier__first_name',
            'session__cash_shift__cashier__last_name',
            'session__cash_shift__cashier__username',
        )
        .annotate(total=Sum('total'), count=Count('id'))
        .order_by('-total')
    )

    # --- helpers ---
    PURPLE   = colors.HexColor('#2D1E5F')
    PINK     = colors.HexColor('#E91E8C')
    LGRAY    = colors.HexColor('#F2F2F2')
    WHITE    = colors.white
    DARK     = colors.HexColor('#1A1A2E')
    DAYS_ES  = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    def fmt(v):
        if v is None:
            return '$0,00'
        s = f'{float(v):,.2f}'
        return '$' + s.replace(',', 'X').replace('.', ',').replace('X', '.')

    def make_table(data, col_widths, header_color=PURPLE):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  header_color),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  WHITE),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('ALIGN',         (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN',         (0, 1), (0, -1),  'LEFT'),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [WHITE, LGRAY]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ]))
        return t

    h2_style = ParagraphStyle(
        'h2', fontName='Helvetica-Bold', fontSize=11, textColor=PURPLE,
        spaceBefore=14, spaceAfter=5,
    )
    caption_style = ParagraphStyle(
        'caption', fontName='Helvetica', fontSize=8, textColor=colors.grey,
        alignment=TA_CENTER,
    )

    # --- build PDF ---
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm,
    )
    W = doc.width
    elements = []

    # Title block
    title_data = [['REPORTE DE VENTAS — CHE GOLOSO']]
    title_tbl = Table(title_data, colWidths=[W])
    title_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), PURPLE),
        ('TEXTCOLOR',     (0, 0), (-1, -1), WHITE),
        ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 15),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(title_tbl)
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f'Período: {start_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}'
        f'  |  Generado: {timezone.localdate().strftime("%d/%m/%Y")}',
        caption_style,
    ))
    elements.append(Spacer(1, 12))

    # --- Resumen ---
    elements.append(Paragraph('Resumen del Período', h2_style))
    total_v = float(stats['total_sales'] or 0)
    dias = (end_date - start_date).days + 1
    resumen_data = [
        ['Indicador', 'Valor'],
        ['Total Vendido',             fmt(stats['total_sales'])],
        ['Cantidad de Transacciones', str(stats['count'] or 0)],
        ['Ticket Promedio',           fmt(stats['avg'])],
        ['Total Descuentos',          fmt(stats['discount'])],
        ['Días analizados',           str(dias)],
        ['Promedio Diario',           fmt(total_v / dias if dias else 0)],
    ]
    elements.append(make_table(resumen_data, [10*cm, 7*cm], PINK))
    elements.append(Spacer(1, 10))

    # --- Métodos de Pago ---
    if by_payment:
        elements.append(Paragraph('Métodos de Pago', h2_style))
        pm_data = [['Método de Pago', 'Transacciones', 'Total', '% del Total']]
        for pm in by_payment:
            pct = (float(pm['total'] or 0) / total_v * 100) if total_v else 0.0
            pm_data.append([
                pm['payment_method__name'],
                str(pm['count']),
                fmt(pm['total']),
                f'{pct:.1f}%',
            ])
        elements.append(make_table(pm_data, [6.5*cm, 3*cm, 4*cm, 3*cm]))
        elements.append(Spacer(1, 10))

    # --- Ventas Diarias ---
    if daily_list:
        elements.append(Paragraph('Ventas Diarias', h2_style))
        daily_data = [['Fecha', 'Día', 'Transacciones', 'Total del Día', 'Acumulado']]
        acumulado = 0.0
        for day in daily_list:
            acumulado += float(day['total'] or 0)
            daily_data.append([
                day['date'].strftime('%d/%m/%Y'),
                DAYS_ES[day['date'].weekday()],
                str(day['count']),
                fmt(day['total']),
                fmt(acumulado),
            ])
        elements.append(make_table(daily_data, [3.5*cm, 2*cm, 3*cm, 4*cm, 4*cm]))
        elements.append(Spacer(1, 10))

    # --- Top Productos ---
    if top_products:
        elements.append(Paragraph('Productos Más Vendidos', h2_style))
        prod_data = [['Producto', 'Categoría', 'Cantidad', 'Ingresos']]
        for item in top_products:
            prod_data.append([
                item['product__name'],
                item['product__category__name'] or 'Sin categoría',
                str(round(float(item['total_qty'] or 0), 3)),
                fmt(item['total_amount']),
            ])
        elements.append(make_table(prod_data, [6.5*cm, 4*cm, 2.5*cm, 3.5*cm]))
        elements.append(Spacer(1, 10))

    # --- Por Cajero ---
    if by_cashier:
        elements.append(Paragraph('Rendimiento por Cajero', h2_style))
        caj_data = [['Cajero', 'Transacciones', 'Total Vendido', 'Ticket Promedio']]
        for c in by_cashier:
            fn = c['session__cash_shift__cashier__first_name'] or ''
            ln = c['session__cash_shift__cashier__last_name'] or ''
            name = f'{fn} {ln}'.strip() or c['session__cash_shift__cashier__username'] or '-'
            avg_caj = float(c['total'] or 0) / c['count'] if c['count'] else 0.0
            caj_data.append([name, str(c['count']), fmt(c['total']), fmt(avg_caj)])
        elements.append(make_table(caj_data, [5*cm, 3.5*cm, 4*cm, 4*cm]))

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_ventas_{start_date}_{end_date}.pdf"'
    )
    return response


# API for real-time stats
@login_required
def api_today_stats(request):
    """API endpoint for real-time today stats."""
    today = timezone.localdate()
    
    transactions = POSTransaction.objects.filter(
        created_at__date=today,
        status='completed'
    )
    
    stats = transactions.aggregate(
        total=Sum('total'),
        count=Count('id')
    )
    
    # By payment method
    payments = POSPayment.objects.filter(
        transaction__created_at__date=today,
        transaction__status='completed'
    ).values('payment_method__name', 'payment_method__color').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Convert Decimals to floats for JSON serialization
    payments_list = [
        {
            'payment_method__name': p['payment_method__name'],
            'payment_method__color': p['payment_method__color'],
            'total': float(p['total'] or 0),
            'count': p['count']
        }
        for p in payments
    ]
    
    return JsonResponse({
        'success': True,
        'total': float(stats['total'] or 0),
        'count': stats['count'] or 0,
        'by_payment': payments_list,
        'timestamp': timezone.now().isoformat()
    })
